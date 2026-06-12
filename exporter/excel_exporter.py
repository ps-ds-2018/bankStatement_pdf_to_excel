from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from extractor.models import (
    Metadata,
    Transaction,
    WarningItem,
)


class ExcelExporter:

    # Severity row fills
    ERROR_FILL   = PatternFill(fill_type="solid", fgColor="F4CCCC")  # red
    WARNING_FILL = PatternFill(fill_type="solid", fgColor="FCE5CD")  # orange
    INFO_FILL    = PatternFill(fill_type="solid", fgColor="FFF2CC")  # yellow

    # Header row style
    HEADER_FILL  = PatternFill(fill_type="solid", fgColor="D9E1F2")  # blue-grey
    HEADER_FONT  = Font(bold=True)

    @staticmethod
    def export(
        output_path: str,
        metadata: Metadata,
        transactions: List[Transaction],
        warnings: List[WarningItem],
    ) -> None:

        output_file = Path(output_path)

        # -----------------------
        # Metadata sheet
        # -----------------------

        metadata_rows = [
            ["Account Holder Name", metadata.account_holder_name],
            ["Account Number",      metadata.account_number],
            ["Statement Period",    metadata.statement_period],
            ["IFSC",                metadata.ifsc],
            ["Branch",              metadata.branch],
            ["Opening Balance",     metadata.opening_balance],
            ["Closing Balance",     metadata.closing_balance],
        ]

        metadata_df = pd.DataFrame(
            metadata_rows,
            columns=["Field", "Value"],
        )

        # -----------------------
        # Transactions sheet
        # -----------------------

        # Collect all headers encountered across all transactions, then
        # reorder them to match the canonical PDF column order:
        #   Date  →  narration/remarks col  →  Debit  →  Credit  →  Balance
        # Any extra columns not in the canonical list are appended at the end.
        all_headers: List[str] = []
        for txn in transactions:
            for header in txn.data.keys():
                if header not in all_headers:
                    all_headers.append(header)

        CANONICAL_ORDER = [
            # Date column (any variant)
            "DATE", "TXN DATE", "TRANSACTION DATE", "VALUE DATE", "VALUE DT",
            # Narration column (any variant)
            "REMARKS", "PARTICULARS", "NARRATION", "DESCRIPTION", "DETAILS",
            # Debit / withdrawal
            "DEBIT", "DEBITS", "WITHDRAWAL", "WITHDRAWALS",
            "WITHDRAWAL AMT", "WITHDRAWAL AMT.", "WITHDRAWALAMT", "WITHDRAWALAMT.",
            # Credit / deposit
            "CREDIT", "CREDITS", "DEPOSIT", "DEPOSITS",
            "DEPOSIT AMT", "DEPOSIT AMT.", "DEPOSITAMT", "DEPOSITAMT.",
            # Balance
            "BALANCE", "CLOSING BALANCE", "CLOSINGBALANCE",
        ]

        def _canonical_rank(header: str) -> int:
            upper = header.upper().strip()
            for rank, key in enumerate(CANONICAL_ORDER):
                if upper == key or key in upper:
                    return rank
            return len(CANONICAL_ORDER)  # unknown columns go last

        all_headers.sort(key=_canonical_rank)

        transaction_rows = [
            {h: txn.data.get(h, "") for h in all_headers}
            for txn in transactions
        ]

        transaction_df = pd.DataFrame(
            transaction_rows,
            columns=all_headers,
        )

        # -----------------------
        # Warnings sheet
        # -----------------------

        warning_rows = [
            {
                "Page":        w.page,
                "Transaction": w.transaction,
                "Issue":       w.issue,
                "Severity":    w.severity,
            }
            for w in warnings
        ]

        warnings_df = pd.DataFrame(warning_rows)

        # -----------------------
        # Write sheets
        # -----------------------

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

            metadata_df.to_excel(
                writer, sheet_name="Metadata", index=False
            )
            transaction_df.to_excel(
                writer, sheet_name="Transactions", index=False
            )
            warnings_df.to_excel(
                writer, sheet_name="Warnings", index=False
            )

        # -----------------------
        # Post-process formatting
        # -----------------------

        wb = load_workbook(output_file)
        ExcelExporter._format_metadata(wb["Metadata"])
        ExcelExporter._format_transactions(wb["Transactions"])
        ExcelExporter._format_warnings(wb["Warnings"])
        wb.save(output_file)

    # --------------------------------------------------
    # Sheet formatters
    # --------------------------------------------------

    @staticmethod
    def _apply_header_style(sheet) -> None:
        """Bold + coloured header row on any sheet."""
        for cell in sheet[1]:
            cell.font = ExcelExporter.HEADER_FONT
            cell.fill = ExcelExporter.HEADER_FILL

    @staticmethod
    def _autofit_columns(sheet, min_width=10, max_width=60) -> None:
        """Set each column width to fit its longest value."""
        for col in sheet.columns:
            width = max(
                len(str(cell.value or "")) for cell in col
            )
            sheet.column_dimensions[
                get_column_letter(col[0].column)
            ].width = min(max(width + 3, min_width), max_width)

    @staticmethod
    def _format_metadata(sheet) -> None:
        ExcelExporter._apply_header_style(sheet)
        ExcelExporter._autofit_columns(sheet)

    @staticmethod
    def _format_transactions(sheet) -> None:
        ExcelExporter._apply_header_style(sheet)
        sheet.freeze_panes = "A2"

        # Wrap text in narration-like columns; right-align numeric ones
        NUMERIC_KEYWORDS = {
            "BALANCE", "DEBIT", "CREDIT",
            "WITHDRAWAL", "DEPOSIT", "AMOUNT",
        }
        NARRATION_KEYWORDS = {
            "NARRATION", "PARTICULARS", "DESCRIPTION", "DETAILS", "REMARKS",
        }

        header_cells = list(sheet[1])
        for cell in header_cells:
            col_name = str(cell.value or "").upper()
            col_letter = get_column_letter(cell.column)

            if any(k in col_name for k in NARRATION_KEYWORDS):
                # Wrap + wider column for narration
                for row_cell in sheet[col_letter]:
                    row_cell.alignment = Alignment(wrap_text=True)
                sheet.column_dimensions[col_letter].width = 45

            elif any(k in col_name for k in NUMERIC_KEYWORDS):
                # Right-align amounts
                for row_cell in sheet[col_letter]:
                    row_cell.alignment = Alignment(horizontal="right")
                sheet.column_dimensions[col_letter].width = 16

            else:
                # Default autofit for date, mode, ref columns
                max_len = max(
                    len(str(c.value or ""))
                    for c in sheet[col_letter]
                )
                sheet.column_dimensions[col_letter].width = min(
                    max(max_len + 3, 10), 40
                )

        # Row height — taller rows for wrapped narration
        for row in sheet.iter_rows(min_row=2):
            sheet.row_dimensions[row[0].row].height = 30

    @staticmethod
    def _format_warnings(sheet) -> None:
        ExcelExporter._apply_header_style(sheet)
        ExcelExporter._autofit_columns(sheet)

        # Find severity column index
        severity_col = next(
            (cell.column for cell in sheet[1] if cell.value == "Severity"),
            None,
        )

        if not severity_col:
            return

        FILL_MAP = {
            "ERROR":   ExcelExporter.ERROR_FILL,
            "WARNING": ExcelExporter.WARNING_FILL,
            "INFO":    ExcelExporter.INFO_FILL,
        }

        for row in range(2, sheet.max_row + 1):
            severity = str(
                sheet.cell(row=row, column=severity_col).value or ""
            ).upper()
            fill = FILL_MAP.get(severity, ExcelExporter.INFO_FILL)
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).fill = fill